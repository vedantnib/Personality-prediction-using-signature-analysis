from openpyxl import *
workbook = load_workbook('C:/Users/Vedant/Desktop/Book2.xlsx')
sheet=workbook['Sheet1']
max_column=sheet.max_column
max_row=sheet.max_row

for i in range(2,872):
	sheet.cell(i,1).value="train_"+str(i-1)
	if i<=438:
		sheet.cell(i,2).value=1
		sheet.cell(i,3).value=0
	else:
		sheet.cell(i,2).value=0
		sheet.cell(i,3).value=1
workbook.save('C:/Users/Vedant/Desktop/Book2.xlsx')